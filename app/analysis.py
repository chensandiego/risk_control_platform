import re
import time
import math
import collections
import io
import zipfile
import tarfile
import os
from celery_app import celery_app
import torch
from torchvision.models.detection import fasterrcnn_resnet50_fpn
from torchvision.transforms import functional as F
from PIL import Image
import pytesseract
from pdfminer.high_level import extract_text
from docx import Document
from openpyxl import load_workbook

from . import crud, schemas, rules_crud
from .database import SessionLocal

# Define sensitive data patterns and their associated risk weights
SENSITIVE_PATTERNS = {
    "email_addresses": {
        "pattern": r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
        "weight": 5,
        "description": "Potential email addresses found."
    },
    "credit_card_numbers": {
        "pattern": r'\b(?:\d[ -]*?){13,16}\b',
        "weight": 10,
        "description": "Potential credit card numbers found."
    },
    "api_keys": {
        "pattern": r"(?:api_key|API_KEY|token|bearer|secret)[\s=:]*['\"]?([a-zA-Z0-9-_\.]{16,})['\"]?",
        "weight": 15,
        "description": "Potential API keys or tokens found."
    },
    "social_security_numbers": {
        "pattern": r'\b\d{3}-\d{2}-\d{4}\b',
        "weight": 20,
        "description": "Potential Social Security Numbers found."
    },
    "private_keys": {
        "pattern": r'-----BEGIN (RSA|DSA|EC|PGP) PRIVATE KEY-----',
        "weight": 25,
        "description": "Potential private cryptographic keys found."
    }
}

# Load a pre-trained object detection model
model = fasterrcnn_resnet50_fpn(pretrained=True)
model.eval()

# COCO class names
COCO_INSTANCE_CATEGORY_NAMES = [
    '__background__', 'person', 'bicycle', 'car', 'motorcycle', 'airplane', 'bus',
    'train', 'truck', 'boat', 'traffic light', 'fire hydrant', 'N/A', 'stop sign',
    'parking meter', 'bench', 'bird', 'cat', 'dog', 'horse', 'sheep', 'cow',
    'elephant', 'bear', 'zebra', 'giraffe', 'N/A', 'backpack', 'umbrella', 'N/A', 'N/A',
    'handbag', 'tie', 'suitcase', 'frisbee', 'skis', 'snowboard', 'sports ball',
    'kite', 'baseball bat', 'baseball glove', 'skateboard', 'surfboard', 'tennis racket',
    'bottle', 'N/A', 'wine glass', 'cup', 'fork', 'knife', 'spoon', 'bowl',
    'banana', 'apple', 'sandwich', 'orange', 'broccoli', 'carrot', 'hot dog', 'pizza',
    'donut', 'cake', 'chair', 'couch', 'potted plant', 'bed', 'N/A', 'dining table',
    'N/A', 'N/A', 'toilet', 'N/A', 'tv', 'laptop', 'mouse', 'remote', 'keyboard', 'cell phone',
    'microwave', 'oven', 'toaster', 'sink', 'refrigerator', 'N/A', 'book',
    'clock', 'vase', 'scissors', 'teddy bear', 'hair drier', 'toothbrush'
]

def extract_text_from_file(contents, mime_type):
    text = ""
    if "text" in mime_type or "json" in mime_type or "xml" in mime_type:
        text = contents.decode('utf-8', errors='ignore')
    elif "pdf" in mime_type:
        text = extract_text(io.BytesIO(contents))
    elif "vnd.openxmlformats-officedocument.wordprocessingml.document" in mime_type:
        doc = Document(io.BytesIO(contents))
        text = "\n".join([para.text for para in doc.paragraphs])
    elif "vnd.openxmlformats-officedocument.spreadsheetml.sheet" in mime_type:
        workbook = load_workbook(filename=io.BytesIO(contents))
        sheet_text = []
        for sheet in workbook.active:
            for row in sheet.iter_rows():
                sheet_text.append(" ".join([str(cell.value) for cell in row]))
        text = "\n".join(sheet_text)
    elif "image" in mime_type:
        image = Image.open(io.BytesIO(contents))
        text = pytesseract.image_to_string(image)
    return text

def analyze_image_content(image_bytes: bytes):
    image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    image_tensor = F.to_tensor(image).unsqueeze(0)

    with torch.no_grad():
        prediction = model(image_tensor)

    detected_objects = []
    for i, score in enumerate(prediction[0]['scores']):
        if score > 0.8: # Confidence threshold
            label_index = prediction[0]['labels'][i].item()
            label = COCO_INSTANCE_CATEGORY_NAMES[label_index]
            detected_objects.append(label)
    return detected_objects

def shannon_entropy(data):
    """Calculate the Shannon entropy of a string."""
    if not data:
        return 0
    entropy = 0
    for x in set(data):
        p_x = float(data.count(x))/len(data)
        if p_x > 0:
            entropy += - p_x*math.log(p_x, 2)
    return entropy

def redact_file(content: bytes, findings: dict) -> bytes:
    content_str = content.decode('utf-8', errors='ignore')
    for category, data in findings.items():
        if "matches" in data:
            for match in data["matches"]:
                content_str = content_str.replace(match, "[REDACTED]")
    return content_str.encode('utf-8')

def quarantine_file(content: bytes, filename: str):
    quarantine_dir = "quarantine"
    if not os.path.exists(quarantine_dir):
        os.makedirs(quarantine_dir)
    with open(os.path.join(quarantine_dir, filename), "wb") as f:
        f.write(content)

@celery_app.task
def analyze_archive_task(content: bytes, content_type: str, filename: str):
    archive_results = {"files": [], "overall_risk_score": 0}
    if "zip" in content_type:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for info in zf.infolist():
                if not info.is_dir():
                    with zf.open(info.filename) as f:
                        file_content = f.read()
                        result = analyze_file_task(file_content, "application/octet-stream", info.filename)
                        archive_results["files"].append(result)
                        archive_results["overall_risk_score"] += result.get("overall_risk_score", 0)
    elif "tar" in content_type:
        with tarfile.open(fileobj=io.BytesIO(content)) as tf:
            for member in tf.getmembers():
                if member.isfile():
                    f = tf.extractfile(member)
                    if f:
                        file_content = f.read()
                        result = analyze_file_task(file_content, "application/octet-stream", member.name)
                        archive_results["files"].append(result)
                        archive_results["overall_risk_score"] += result.get("overall_risk_score", 0)
    return archive_results

@celery_app.task
def analyze_file_task(content: bytes, content_type: str, filename: str):
    if "zip" in content_type or "tar" in content_type:
        return analyze_archive_task(content, content_type, filename)

    time.sleep(5)
    total_risk_score = 0
    findings = {}

    if content_type.startswith("image/"):
        detected_objects = analyze_image_content(content)
        if detected_objects:
            findings["image_objects"] = {
                "count": len(detected_objects),
                "matches": detected_objects,
                "description": "Objects detected in image.",
                "risk_contribution": len(detected_objects) * 10
            }
            total_risk_score += len(detected_objects) * 10

    content_str = extract_text_from_file(content, content_type)

    # Regex-based scanning
    for category, data in SENSITIVE_PATTERNS.items():
        matches = re.findall(data["pattern"], content_str)
        if matches:
            findings[category] = {
                "count": len(matches),
                "matches": matches,
                "description": data["description"],
                "risk_contribution": len(matches) * data["weight"]
            }
            total_risk_score += len(matches) * data["weight"]

    # Custom rule-based scanning
    db = SessionLocal()
    custom_rules = rules_crud.get_rules(db)
    db.close()

    for rule in custom_rules:
        matches = re.findall(rule.pattern, content_str)
        if matches:
            category = f"custom_rule_{rule.id}"
            findings[category] = {
                "count": len(matches),
                "matches": matches,
                "description": rule.description,
                "risk_contribution": len(matches) * 10  # Default weight for custom rules
            }
            total_risk_score += len(matches) * 10

    # Entropy-based secret detection
    high_entropy_strings = []
    for word in content_str.split():
        if len(word) > 20: # Only check longer strings
            entropy = shannon_entropy(word)
            if entropy > 4.5: # High entropy threshold
                high_entropy_strings.append(word)

    if high_entropy_strings:
        findings["high_entropy_secrets"] = {
            "count": len(high_entropy_strings),
            "matches": high_entropy_strings,
            "description": "Potential secrets detected based on high entropy.",
            "risk_contribution": len(high_entropy_strings) * 30 # High weight for entropy-based findings
        }
        total_risk_score += len(high_entropy_strings) * 30

    anomalies = []
    if len(content_str) > 100000:
        anomalies.append("File size is unusually large (over 100KB).")
        total_risk_score += 50

    if anomalies:
        findings["anomalies"] = {
            "count": len(anomalies),
            "matches": anomalies,
            "description": "Potential anomalies detected.",
            "risk_contribution": 50
        }

    if total_risk_score > 100:
        quarantine_file(content, filename)

    summary = generate_risk_summary(total_risk_score, findings)

    analysis_data = {
        "overall_risk_score": total_risk_score,
        "detailed_findings": findings,
        "summary": summary,
        "filename": filename,
        "content_type": content_type
    }

    # Save to MongoDB
    result_to_save = schemas.AnalysisResultCreate(
        filename=filename,
        content_type=content_type,
        risk_score=analysis_data['overall_risk_score'],
        findings=analysis_data['detailed_findings']
    )
    saved_result = crud.create_analysis_result(result=result_to_save)

    # Add the MongoDB ID to the returned data for caching and frontend display
    analysis_data["_id"] = str(saved_result["_id"])
    analysis_data["id"] = str(saved_result["_id"])

    return analysis_data

def generate_risk_summary(score: int, findings: dict):
    summary_lines = [f"Overall Risk Score: {score}"]
    
    if score == 0:
        summary_lines.append("No significant risks detected.")
    elif score < 50:
        summary_lines.append("Low risk: Minor issues found.")
    elif score < 150:
        summary_lines.append("Medium risk: Some sensitive data or anomalies detected.")
    else:
        summary_lines.append("High risk: Significant sensitive data or critical anomalies detected. Immediate review recommended.")

    if findings:
        summary_lines.append("\nDetailed Findings:")
        for category, data in findings.items():
            summary_lines.append(f"- {data['description']} (Count: {data['count']}, Risk Contribution: {data['risk_contribution']})")
    
    return "\n".join(summary_lines)

if __name__ == "__main__":
    sample_content = b"my api key is: aS1k2j3h4g5f6d7s8a9p0o1i2u3y4t5r6e7w8q9z0x. and some other text"
    # To test the task, you would typically call .delay() or .apply_async()
    # For simple local testing, you can call the function directly:
    # result = analyze_file_task(sample_content, "text/plain", "sample.txt")
    # print(result["summary"])
    pass

